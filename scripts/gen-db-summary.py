#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parse prisma/schema.prisma -> docs/DATABASE_SUMMARY.xlsx (multi-sheet overview + collection checklist)."""
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/Jun/Documents/CRM-D&Z")
SCHEMA = ROOT / "prisma/schema.prisma"
OUT = ROOT / "docs/DATABASE_SUMMARY.xlsx"

src = SCHEMA.read_text(encoding="utf-8").splitlines()
models = []
enums = []
cur = None
pending_comment = ""
i = 0
RE_MODEL = re.compile(r"^model\s+(\w+)\s*\{")
RE_ENUM = re.compile(r"^enum\s+(\w+)\s*\{")
RE_FIELD = re.compile(r"^(\w+)\s+([\w\[\]]+)(\??)(\s+@[^/]*)?(\s*//\s*(.*))?$")
RE_DEFAULT = re.compile(r"@default\(([^)]*)\)")
while i < len(src):
    line = src[i].strip()
    m = RE_MODEL.match(line)
    e = RE_ENUM.match(line)
    if m:
        cur = {"name": m.group(1), "comment": pending_comment, "fields": []}
        pending_comment = ""
        i += 1
        while i < len(src) and "}" not in src[i]:
            raw = src[i].strip()
            if raw and not raw.startswith("//"):
                fm = RE_FIELD.match(raw)
                if fm:
                    fname, ftype, q, attrs, _, fcomment = fm.groups()
                    default = ""
                    if attrs:
                        dm = RE_DEFAULT.search(attrs)
                        if dm:
                            default = dm.group(1)
                    cur["fields"].append({"name": fname, "type": ftype, "required": not q, "default": default, "comment": (fcomment or "").strip()})
            i += 1
        models.append(cur)
        cur = None
    elif e:
        cur = {"name": e.group(1), "values": []}
        i += 1
        while i < len(src) and "}" not in src[i]:
            raw = src[i].strip()
            if raw and not raw.startswith("//"):
                cur["values"].append(raw)
            i += 1
        enums.append(cur)
        cur = None
    else:
        if line.startswith("//") and not line.startswith("///"):
            pending_comment = line.lstrip("/").strip()
        i += 1

print("parsed", len(models), "models,", len(enums), "enums")

wb = Workbook()
HEAD = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
BAND = PatternFill("solid", fgColor="F3F4F6")
WRAP = Alignment(vertical="top", wrap_text=True)

def sheet_headers(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEAD; cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

def autoband(ws, start=2):
    for r in range(start, ws.max_row + 1):
        if (r - start) % 2 == 1:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = BAND

CN = {
    "Organisation":"组织（经销商公司）","Branch":"门店/分支","User":"用户（员工）","Customer":"客户","CustomerAddress":"客户地址",
    "CustomerConsent":"客户同意（PDPA）","Motorcycle":"摩托车（车辆档案）","ServiceHistory":"服务历史","Product":"产品/配件",
    "InventoryLocation":"库存位置","Supplier":"供应商","PurchaseOrder":"采购订单","StockMovement":"库存流水","Inventory":"库存",
    "Lead":"销售线索","LeadSource":"线索来源","LeadStage":"线索阶段","LeadActivity":"线索活动","Task":"跟进任务","TestRide":"试驾预约",
    "Booking":"服务预约","AppointmentSlot":"预约时段","ServiceJob":"服务工单","JobStatusHistory":"工单状态历史","ServiceType":"服务类型",
    "ServicePackage":"服务套餐","PackageItem":"套餐项目","ServiceJobItem":"工单服务项","ServiceJobPart":"工单配件","InspectionFinding":"检查发现",
    "Approval":"审批（增项）","Invoice":"发票","Payment":"付款","Checklist":"检查清单","Campaign":"营销活动","MarketingAsset":"营销素材（海报）",
    "Message":"消息记录","MessageTemplate":"消息模板","Notification":"通知","LoyaltyAccount":"忠诚度账户","LoyaltyTier":"忠诚度等级",
    "LoyaltyTransaction":"积分流水","Reward":"奖励","RewardRedemption":"奖励兑换","Referral":"推荐","Review":"评价",
    "AutomationRule":"自动化规则","AutomationExecution":"自动化执行","AuditLog":"审计日志","IntegrationConfig":"集成配置",
    "Attachment":"附件","RoleConfig":"角色配置","Permission":"权限","Reminder":"保养提醒","ReturnVisit":"返修/回访","InsuranceQuote":"保险报价",
    "FinancingQuote":"分期报价","Warranty":"保修","Recall":"召回","ServiceCatalog":"服务目录","KpiSnapshot":"KPI 快照","VehicleHandover":"交车记录",
}

ws = wb.active; ws.title = "1. 数据模型总览"
sheet_headers(ws, ["表名", "中文说明", "字段数", "核心用途"], [28, 34, 8, 60])
for r_i, m in enumerate(models, 2):
    ws.cell(row=r_i, column=1, value=m["name"])
    ws.cell(row=r_i, column=2, value=CN.get(m["name"], ""))
    ws.cell(row=r_i, column=3, value=len(m["fields"]))
    ws.cell(row=r_i, column=4, value=m["comment"])
autoband(ws)

ws = wb.create_sheet("2. 字段明细")
sheet_headers(ws, ["表", "字段", "类型", "必填", "默认值", "说明"], [26, 24, 14, 8, 22, 46])
r = 2
for m in models:
    for f in m["fields"]:
        ws.cell(row=r, column=1, value=m["name"])
        ws.cell(row=r, column=2, value=f["name"])
        ws.cell(row=r, column=3, value=f["type"])
        ws.cell(row=r, column=4, value="Y" if f["required"] else "N")
        ws.cell(row=r, column=5, value=f["default"])
        ws.cell(row=r, column=6, value=f["comment"])
        for c in range(1, 7): ws.cell(row=r, column=c).alignment = WRAP
        r += 1
autoband(ws)

ws = wb.create_sheet("3. 数据收集清单")
sheet_headers(ws, ["业务域", "表", "要收集的资料", "收集来源/渠道", "必填", "示例 / 说明"], [18, 24, 46, 26, 8, 50])
COLLECT = [
 ("组织信息","Organisation","公司名称、地址、联系电话、税号/注册号、Logo","经销商管理层","Y","门店注册时的基础资料"),
 ("组织信息","Branch","门店名、城市、地址、电话、营业时间、是否为总店","门店经理","Y","营业时间格式如 09:00-18:00"),
 ("员工","User","姓名、角色、电话、邮箱、登录密码、MFA 密钥","HR / 店长","Y","角色见枚举 Role（16 种）"),
 ("客户","Customer","姓名、电话、邮箱、地址、来源渠道、加入日期","前台登记 / 线上表单","Y","电话是唯一标识，用于 WhatsApp 触达"),
 ("客户","CustomerAddress","客户配送/联络地址","前台 / 客户自填","N","多地址支持"),
 ("客户","CustomerConsent","PDPA 营销同意状态、同意时间","客户签署","Y","马来西亚 PDPA 合规必需"),
 ("车辆","Motorcycle","品牌、型号、车牌、VIN、年份、类型、颜色、当前里程、上次/下次保养里程","客户提供 + 技师录入","Y","类型枚举 12 类"),
 ("车辆","ServiceHistory","每次服务的日期、里程、项目、费用","工单完成自动写入","Y","由 ServiceJob 完成后生成"),
 ("产品库存","Product","SKU、名称、类别、品牌、成本价、售价、条形码、厂商料号、安全库存","采购 / 库存经理","Y","成本价敏感，单独权限"),
 ("产品库存","Supplier","供应商名、联系人、电话、地址、账期","采购","Y","用于采购订单与再订货"),
 ("产品库存","PurchaseOrder","PO 号、供应商、商品明细、数量、成本、状态","采购流程","Y","收货后自动扣减库存"),
 ("产品库存","StockMovement","出入库流水（类型、数量、原因、操作人）","系统自动","Y","盘点/调拨/收货/销售出库"),
 ("销售线索","Lead","客户名、电话、邮箱、意向车型、来源、阶段、预估价值、下次跟进","网站表单 / 前台 / 销售","Y","公开网站咨询/试驾自动生成"),
 ("销售线索","LeadStage","线索阶段名称、顺序、胜率","销售主管","Y","如 新线索->已联系->试驾->报价->成交"),
 ("销售线索","TestRide","试驾客户、日期时段、车型、门店、结果","线上表单 / 前台","Y","可预约试驾"),
 ("服务预约","Booking","客户、车辆、服务类型、日期、时段、门店、来源","Rider App / 前台","Y","槽位防超卖"),
 ("服务预约","AppointmentSlot","日期、时段、最大接单数、是否假日","店长配置","Y","假日可设 0 或关闭"),
 ("服务工单","ServiceJob","工单号、客户、车辆、里程、状态、技师、套餐、开始/完成时间","前台 / 技师","Y","9 状态流转"),
 ("服务工单","ServicePackage","套餐名、档位、价格、描述、是否最佳值、包含项目/赠品","服务主管","Y","GOOD/BETTER/BEST 套餐体系"),
 ("服务工单","ServiceType","服务类型名、分类、价格参考","服务主管","Y","如 普通保养/大保养/维修"),
 ("营销","Campaign","活动名、类型、折扣%、开始/结束日期、状态","市场部","Y","促销可关联预约转化归因"),
 ("营销","MarketingAsset","海报/素材标题、类型、尺寸、色调、文件","市场部 + AI 生成","Y","AI 海报生成功能"),
 ("营销","MessageTemplate","模板名、渠道、标题、正文（占位符）、状态","市场部","N","支持 {name}/{bike}/{link} 占位符"),
 ("营销","Referral","推荐人、被推荐人、代码、状态、奖励","客户推荐","N","推荐码追踪"),
 ("忠诚度","LoyaltyAccount","客户、会员号、积分余额、累计赚取/兑换、等级","系统 + 前台","Y","积分体系（Earn/Adjust/Redeem）"),
 ("忠诚度","LoyaltyTier","等级名、门槛、权益描述","店长配置","Y","如 银/金/铂金"),
 ("忠诚度","Reward","奖励名、所需积分、是否启用","店长配置","Y","兑换礼品/折扣"),
 ("财务","Invoice","发票号、工单、合计、毛利、状态、开票日期","系统生成","Y","工单完成自动出票"),
 ("财务","Payment","付款方式、金额、状态、时间","收银","Y","支持分期/全款"),
 ("评价","Review","客户、门店、评分、评论、来源、状态、回复","Rider App / 线上","N","评分运营"),
 ("系统","IntegrationConfig","渠道配置（WhatsApp/OpenAI/存储/支付）、启用状态","IT 管理员","N","Provider 抽象"),
 ("系统","AutomationRule","触发事件、动作、条件、是否启用","店长 / 管理员","N","10 触发 x 5 动作"),
 ("系统","AuditLog","操作、实体、前后值、用户、IP、时间","系统自动","Y","敏感操作留痕"),
]
for r_i, row in enumerate(COLLECT, 2):
    for c_i, v in enumerate(row, 1):
        cell = ws.cell(row=r_i, column=c_i, value=v)
        cell.alignment = WRAP
autoband(ws)

ws = wb.create_sheet("4. 枚举与状态")
sheet_headers(ws, ["枚举", "值", "说明"], [30, 60, 40])
EN_CN = {
    "Role":"用户角色（16 种）","JobStatus":"工单状态（9 值）","BookingStatus":"预约状态","LeadStatus":"线索状态",
    "InvoiceStatus":"发票状态","PaymentStatus":"付款状态","CampaignStatus":"活动状态","CampaignType":"活动类型",
    "MessageChannel":"消息渠道","MessageStatus":"消息状态","ReferralStatus":"推荐状态","LoyaltyTxType":"积分流水类型",
    "AttachmentKind":"附件类型","UserStatus":"用户状态","NotificationType":"通知类型","MotorcycleType":"车型类型",
    "ProductCategory":"产品类别","InventoryMovementType":"库存流水类型","PurchaseOrderStatus":"采购单状态",
    "ChecklistItemResult":"检查项结果","FindingSeverity":"发现严重度","ApprovalStatus":"审批状态","ServiceTypeCategory":"服务类型分类",
}
r = 2
for e in enums:
    for v in e["values"]:
        ws.cell(row=r, column=1, value=e["name"])
        ws.cell(row=r, column=2, value=v)
        ws.cell(row=r, column=3, value=EN_CN.get(e["name"], ""))
        r += 1
autoband(ws)

ws = wb.create_sheet("5. 表关系")
sheet_headers(ws, ["模型", "关联字段", "目标模型", "关系"], [26, 24, 26, 8])
r = 2
for m in models:
    for f in m["fields"]:
        if f["type"].endswith("]") or (f["name"].endswith("Id") and f["type"] == "String"):
            target = f["type"].rstrip("[]") if f["type"].endswith("]") else f["name"][:-2]
            rel = "N->1" if f["type"].endswith("]") else "1->1"
            ws.cell(row=r, column=1, value=m["name"])
            ws.cell(row=r, column=2, value=f["name"])
            ws.cell(row=r, column=3, value=target)
            ws.cell(row=r, column=4, value=rel)
            r += 1
autoband(ws)

ws = wb.create_sheet("6. 统计速览")
sheet_headers(ws, ["指标", "数值", "说明"], [30, 18, 50])
stats = [
    ("模型总数", str(len(models)), "Prisma schema 中的数据表"),
    ("枚举总数", str(len(enums)), "状态/类型枚举"),
    ("字段总数", str(sum(len(m["fields"]) for m in models)), "全部字段（含关系字段）"),
    ("核心业务表", "24", "组织/员工/客户/车辆/产品/线索/工单/预约/营销/财务等"),
    ("扩展表", str(len(models) - 24), "忠诚度/自动化/审计/集成/附件等"),
]
for r_i, row in enumerate(stats, 2):
    for c_i, v in enumerate(row, 1):
        ws.cell(row=r_i, column=c_i, value=v)
autoband(ws)

wb.save(OUT)
print("saved", OUT)
