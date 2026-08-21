#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Dealer onboarding data collection form — standalone Excel for the dealer to fill.
Columns: domain / table / what to collect / source / required / example / status / owner / date. """
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/Jun/Documents/CRM-D&Z/docs/DEALER_ONBOARDING_COLLECTION_FORM.xlsx"

COLLECT = [
 ("组织信息","Organisation","公司名称、地址、联系电话、税号/注册号、Logo","经销商管理层","Y","门店注册时的基础资料"),
 ("组织信息","Branch","门店名、城市、地址、电话、营业时间、是否为总店","门店经理","Y","营业时间格式如 09:00-18:00"),
 ("员工","User","姓名、角色、电话、邮箱、登录账号、MFA","HR / 店长","Y","角色：老板/经理/技师/前台/销售等"),
 ("客户","Customer","姓名、电话、邮箱、地址、来源、加入日期","前台登记 / 线上表单","Y","电话唯一，用于 WhatsApp 触达"),
 ("客户","CustomerConsent","PDPA 营销同意状态、同意时间","客户签署","Y","马来西亚 PDPA 合规必需"),
 ("车辆","Motorcycle","品牌、型号、车牌、VIN、年份、类型、颜色、当前里程、保养里程","客户提供 + 技师录入","Y","车型类型：踏板/街车/跑车/巡航等"),
 ("产品库存","Product","SKU、名称、类别、品牌、成本价、售价、条形码、安全库存","采购 / 库存经理","Y","成本价敏感，单独权限"),
 ("产品库存","Supplier","供应商名、联系人、电话、地址、账期","采购","Y","用于采购订单与再订货"),
 ("产品库存","PurchaseOrder","PO 号、供应商、商品明细、数量、成本、状态","采购流程","Y","收货后自动扣减库存"),
 ("销售线索","Lead","客户名、电话、意向车型、来源、阶段、预估价值","前台 / 销售","Y","公开网站咨询自动生成"),
 ("销售线索","LeadStage","线索阶段名称、顺序、胜率","销售主管","Y","如 新线索→已联系→试驾→报价→成交"),
 ("服务预约","Booking","客户、车辆、服务类型、日期、时段、门店","前台 / Rider App","Y","槽位防超卖"),
 ("服务预约","AppointmentSlot","日期、时段、最大接单数、是否假日","店长配置","Y","假日可设 0 或关闭"),
 ("服务工单","ServiceJob","工单号、客户、车辆、里程、状态、技师、套餐","前台 / 技师","Y","9 状态流转"),
 ("服务工单","ServicePackage","套餐名、档位、价格、包含项目/赠品","服务主管","Y","GOOD/BETTER/BEST 套餐体系"),
 ("服务工单","ServiceType","服务类型名、分类、价格参考","服务主管","Y","如 普通保养/大保养/维修"),
 ("营销","Campaign","活动名、类型、折扣%、开始/结束日期、状态","市场部","Y","促销可关联预约转化归因"),
 ("营销","MessageTemplate","模板名、渠道、标题、正文（占位符）","市场部","N","支持 {name}/{bike}/{link} 占位符"),
 ("忠诚度","LoyaltyTier","等级名、门槛、权益描述","店长配置","Y","如 银/金/铂金"),
 ("忠诚度","Reward","奖励名、所需积分、是否启用","店长配置","Y","兑换礼品/折扣"),
 ("财务","Payment","付款方式、金额、状态、时间","收银","Y","支持分期/全款"),
 ("系统","IntegrationConfig","渠道配置（WhatsApp/AI/存储/支付）、启用状态","IT 管理员","N","Provider 抽象，生产接真实服务"),
 ("系统","AutomationRule","触发事件、动作、条件、是否启用","店长 / 管理员","N","10 触发 x 5 动作"),
]

wb = Workbook()
# ---------- Sheet 0: instructions ----------
ws = wb.active; ws.title = "填写指引"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 100
title = ws.cell(row=1, column=1, value="D&Z 平台 · 经销商入驻资料收集表")
title.font = Font(bold=True, size=18)
title.fill = PatternFill("solid", fgColor="1F2937")
title.font = Font(color="FFFFFF", bold=True, size=16)
ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40
guide = [
    "",
    "本表用于收集 D&Z 平台上线前需要经销商提供/整理的基础数据。",
    "",
    "【填写指引】",
    "1. 「必填」= Y 的资料是系统正常运行所必需，请优先收集；N 可上线后补充。",
    "2. 「状态」列请填写：未开始 / 进行中 / 已完成。",
    "3. 「负责人」列填写该资料的实际负责人姓名。",
    "4. 「日期」列填写完成日期（YYYY-MM-DD）。",
    "5. 有疑问的字段可在「备注」列注明。",
    "",
    "【推荐顺序】",
    "① 组织信息 + 员工（账号能登录） → ② 客户 + 车辆（可开单） → ③ 产品 + 供应商（可库存）",
    "→ ④ 服务套餐/类型（可报价） → ⑤ 营销/忠诚度（可运营） → ⑥ 系统集成（可触达客户）",
    "",
    "完整数据库结构见 docs/DATABASE_SUMMARY.xlsx（6 个 sheet）。",
]
for r_i, t in enumerate(guide, 2):
    ws.cell(row=r_i, column=1, value=t)
ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

# ---------- Sheet 1: collection checklist ----------
ws = wb.create_sheet("资料收集清单")
headers = ["业务域", "表", "要收集的资料", "收集来源/渠道", "必填", "示例 / 说明", "状态", "负责人", "完成日期", "备注"]
widths = [16, 20, 44, 22, 8, 40, 12, 12, 14, 18]
HEAD = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
DOMAIN_FILL = PatternFill("solid", fgColor="E8ECF2")
DOMAIN_FONT = Font(bold=True, size=11, color="1F2937")
BAND = PatternFill("solid", fgColor="F6F7F9")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = HEAD; cell.font = HEAD_FONT; cell.alignment = CENTER; cell.border = BORDER
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{len(COLLECT)+1}"

r = 2
prev_domain = None
for row in COLLECT:
    domain = row[0]
    if domain != prev_domain:
        # domain section header row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        dc = ws.cell(row=r, column=1, value="【" + domain + "】")
        dc.fill = DOMAIN_FILL; dc.font = DOMAIN_FONT
        dc.alignment = Alignment(vertical="center")
        for c in range(1, 11): ws.cell(row=r, column=c).fill = DOMAIN_FILL; ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 20
        r += 1
        prev_domain = domain
    vals = list(row) + ["未开始", "", "", ""]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.alignment = CENTER if c in (5, 7) else WRAP
    # required highlight
    if row[4] == "Y":
        ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor="FEF3C7")
        ws.cell(row=r, column=5).font = Font(bold=True, color="B45309")
    if (r % 2) == 0:
        for c in range(1, 11):
            if ws.cell(row=r, column=c).fill.fgColor.rgb in (None, "00000000"):
                ws.cell(row=r, column=c).fill = BAND
    r += 1

# summary footer
r += 1
total = sum(1 for x in COLLECT if x[4] == "Y")
ws.cell(row=r, column=1, value=f"合计 {len(COLLECT)} 项，其中必填 {total} 项").font = Font(bold=True, color="B45309")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

wb.save(OUT)
print("saved", OUT, "| rows:", len(COLLECT))
